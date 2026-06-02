# Importa o histórico de alugueres (2024+2025) do Excel diretamente para o Supabase
# via API REST, usando a chave de serviço (lida do ambiente SUPABASE_SERVICE_KEY).
# Não passa pela consola (acentos ficam corretos).
import openpyxl, unicodedata, os, json, urllib.request

URL = 'https://lykfbclxsyazerffcpta.supabase.co'
KEY = os.environ.get('SUPABASE_SERVICE_KEY', '').strip()
XLSX = r'C:\Users\andre\Documents\Antigravity\Produtividade\Comissoes Actualizadas - Abril  2026.xlsx'
ANOS = ('2026',)

if not KEY:
    raise SystemExit('Falta SUPABASE_SERVICE_KEY no ambiente.')

MESES = {'janeiro':1,'fevereiro':2,'marco':3,'abril':4,'maio':5,'junho':6,
         'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

def sa(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def mes_ano(nome):
    t = sa(nome).lower().strip().split()
    if len(t) >= 2 and t[0] in MESES and t[-1] in ANOS:
        return int(t[-1]), MESES[t[0]]
    return None

def norm_met(s):
    t = sa(s).lower().strip()
    if not t: return None
    ht = 'transfer' in t or t == 'tb' or 'tb ' in t or ' tb' in t
    hn = 'numerar' in t or 'dinheiro' in t
    if 'mbway' in t or 'mb way' in t: return 'MBway'
    if 'cheque' in t: return 'Cheque'
    if ht and hn: return 'Transferência/Numerário'
    if ht: return 'Transferência'
    if hn: return 'Numerário'
    return None

def post(tabela, linhas, prefer, params=''):
    if not linhas: return
    req = urllib.request.Request(
        f'{URL}/rest/v1/{tabela}{params}',
        data=json.dumps(linhas).encode('utf-8'),
        headers={
            'apikey': KEY,
            'Authorization': f'Bearer {KEY}',
            'Content-Type': 'application/json',
            'Prefer': prefer,
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status
    except urllib.error.HTTPError as e:
        corpo = e.read().decode('utf-8', 'replace')
        raise SystemExit(f'ERRO {e.code} em {tabela}: {corpo[:500]}')

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
alvo = [n for n in wb.sheetnames if mes_ano(n)]

alug = []
nomes = set()
for nome in alvo:
    ano, mes = mes_ano(nome)
    data = f'{ano}-{mes:02d}-01'
    rows = list(wb[nome].iter_rows(values_only=True))
    hr = None
    for i, r in enumerate(rows):
        cells = [sa(c).lower().strip() if c is not None else '' for c in r]
        if 'name' in cells:
            hr = i; header = cells; break
    ci_name = header.index('name')
    ci_val = next((j for j,h in enumerate(header) if 'valor' in h), None)
    ci_met = next((j for j,h in enumerate(header) if 'pagamento' in h), None)
    for r in rows[hr+1:]:
        nm = r[ci_name] if ci_name < len(r) else None
        if nm is None or str(nm).strip() == '': break
        if 'total' in sa(nm).lower(): continue
        val = r[ci_val] if (ci_val is not None and ci_val < len(r)) else None
        try: v = float(str(val).replace(',', '.')) if val not in (None, '') else None
        except: v = None
        if v is None or v <= 0: continue
        cn = str(nm).strip()
        nomes.add(cn)
        alug.append({
            'cliente_nome': cn,
            'valor': v,
            'metodo_pagamento': norm_met(r[ci_met]) if (ci_met is not None and ci_met < len(r)) else None,
            'nacional': True,
            'data_entrega': data,
            'criado_por_nome': 'Importado (Excel)',
        })

# Clientes (ignora duplicados pelo nome único)
clientes = [{'nome': n, 'pais': 'Portugal', 'nacional': True} for n in sorted(nomes)]
post('clientes', clientes, 'resolution=ignore-duplicates,return=minimal', '?on_conflict=nome')
print('Clientes enviados:', len(clientes))

# Alugueres em lotes de 300
total = 0
for i in range(0, len(alug), 300):
    lote = alug[i:i+300]
    post('alugueres', lote, 'return=minimal')
    total += len(lote)
    print('Alugueres inseridos:', total)

print('CONCLUÍDO. Total alugueres:', len(alug), '| clientes:', len(clientes))
